from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

from .config import CITY_MAP, SHEET_REQ1, SHEET_REQ2, ANCHORS
from .excel_writer import write_df
from .transforms import (
    clean_columns, filter_by_city, filter_by_canal, filter_presencias_by_group_city,
    transform_edf_final
)

def read_csv(uploaded_file) -> pd.DataFrame:
    """Lee CSV desde Streamlit uploader."""
    return clean_columns(pd.read_csv(uploaded_file))

def build_workbook_bytes(
    template_xlsx_file,
    manejantes_csv,
    presencias_csv,
    sovi_csv,
    edf_csv,
    secos_csv,
) -> bytes:
    """Core: construye el Excel final y lo devuelve como bytes para descarga."""

    # 1) Leer data
    df_manejantes = read_csv(manejantes_csv)
    df_presencias = read_csv(presencias_csv)
    df_sovi = read_csv(sovi_csv)
    df_edf = read_csv(edf_csv)
    df_secos = read_csv(secos_csv)

    # 2) Cargar template
    template_bytes = template_xlsx_file.getvalue()
    wb = load_workbook(BytesIO(template_bytes))

    # 3) Base REQ
    req_cfg = ANCHORS["REQ"]
    write_df(wb[SHEET_REQ1], df_manejantes, req_cfg["cell"], req_cfg["headers"])
    write_df(wb[SHEET_REQ2], df_presencias, req_cfg["cell"], req_cfg["headers"])

    # 4) TT / C&B (desde presencias)
    ttcb_cfg = ANCHORS["TT_CB"]
    for code, cities in CITY_MAP.items():
        tt = f"{code}-TT"
        if tt in wb.sheetnames:
            d = filter_presencias_by_group_city(df_presencias, "TRADICIONAL", cities)
            write_df(wb[tt], d, ttcb_cfg["cell"], ttcb_cfg["headers"])

        cb = f"{code}-C&B"
        if cb in wb.sheetnames:
            d = filter_presencias_by_group_city(df_presencias, "COMER & BEBER", cities)
            write_df(wb[cb], d, ttcb_cfg["cell"], ttcb_cfg["headers"])

    # 5) VAP (Kiosko + LP/EA)
    vap_cfg = ANCHORS["VAP"]
    for code, sheet in [("LP", "LP-VAP"), ("EA", "EA-VAP")]:
        if sheet in wb.sheetnames:
            d = filter_by_canal(df_presencias, "Kiosko")
            d = filter_by_city(d, CITY_MAP[code])
            write_df(wb[sheet], d, vap_cfg["cell"], vap_cfg["headers"])

    # 6) SOVI (headers)
    sovi_cfg = ANCHORS["SOVI"]
    for code, sheet in [("LP", "SOVI LP"), ("EA", "SOVI EA"), ("OR", "SOVI OR")]:
        if sheet in wb.sheetnames:
            d = filter_by_city(df_sovi, CITY_MAP[code])
            write_df(wb[sheet], d, sovi_cfg["cell"], sovi_cfg["headers"])

    # 7) EDF (transform + ciudad) (headers)
    edf_cfg = ANCHORS["EDF"]
    df_edf_t = transform_edf_final(df_edf)
    for code, cities in CITY_MAP.items():
        sheet = f"{code}-EDF"
        if sheet in wb.sheetnames:
            d = filter_by_city(df_edf_t, cities)
            write_df(wb[sheet], d, edf_cfg["cell"], edf_cfg["headers"])

    # 8) SECOS (headers)
    secos_cfg = ANCHORS["SECOS"]
    for code, sheet in [("LP", "SECOS LP"), ("EA", "SECOS EA"), ("CBBA", "SECOS CBBA")]:
        if sheet in wb.sheetnames:
            d = filter_by_city(df_secos, CITY_MAP[code])
            write_df(wb[sheet], d, secos_cfg["cell"], secos_cfg["headers"])

    # 9) Exportar a bytes
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

