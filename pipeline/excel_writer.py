import pandas as pd
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def clear_block(ws, start_row, start_col, nrows, ncols):
    """Limpia solo el bloque que vamos a sobrescribir (no toda la hoja)."""
    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            ws.cell(row=r, column=c).value = None

def write_df(ws, df: pd.DataFrame, anchor_cell="A2", include_headers=False):
    """Escribe un DF en una hoja, preservando estilos del template (solo cambia valores)."""
    col_letter, row = coordinate_from_string(anchor_cell)
    start_col = column_index_from_string(col_letter)
    start_row = row

    nrows = len(df) + (1 if include_headers else 0) + 10
    ncols = len(df.columns) + 5
    clear_block(ws, start_row, start_col, nrows, ncols)

    r0 = start_row
    if include_headers:
        for j, col in enumerate(df.columns):
            ws.cell(row=r0, column=start_col + j).value = str(col)
        r0 += 1

    for i in range(len(df)):
        row_vals = df.iloc[i].tolist()
        for j, val in enumerate(row_vals):
            ws.cell(row=r0 + i, column=start_col + j).value = None if pd.isna(val) else val

